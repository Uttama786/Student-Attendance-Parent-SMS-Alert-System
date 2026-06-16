"""
blueprints/reports/routes.py — Reporting and data export.

Routes:
  GET /reports/daily              — Daily attendance summary (date filter)
  GET /reports/student/<id>       — Per-student attendance history
  GET /reports/export/excel       — Download Excel workbook
  GET /reports/export/pdf         — Download PDF report
  GET /reports/faculty            — Faculty management (admin only)
  POST /reports/faculty/add       — Add faculty (admin only)
  POST /reports/faculty/<id>/delete — Delete faculty (admin only)
"""
import io
from datetime import date, datetime

from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, send_file, abort,
)
from flask_login import login_required, current_user
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SelectField
from wtforms.validators import DataRequired, Length, EqualTo

from extensions import db
from models import Student, Attendance, SMSLog, Faculty, AttendanceStatus

reports_bp = Blueprint("reports", __name__, template_folder="../../templates/reports")


# ─────────────────────────────── Forms ───────────────────────────────────────

class AddFacultyForm(FlaskForm):
    name = StringField("Full Name", validators=[DataRequired(), Length(max=120)])
    username = StringField("Username", validators=[DataRequired(), Length(max=80)])
    role = SelectField("Role", choices=[("faculty", "Faculty"), ("admin", "Admin")])
    password = PasswordField("Password", validators=[DataRequired(), Length(min=6)])
    confirm = PasswordField(
        "Confirm Password",
        validators=[DataRequired(), EqualTo("password", message="Passwords must match")],
    )


# ─────────────────────────────── Routes ──────────────────────────────────────

@reports_bp.route("/daily")
@login_required
def daily():
    """Daily attendance summary filtered by date."""
    date_str = request.args.get("date", str(date.today()))
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        report_date = date.today()

    records = (
        db.session.query(Attendance, Student)
        .join(Student, Attendance.student_id == Student.id)
        .filter(Attendance.date == report_date)
        .order_by(Student.department, Student.name)
        .all()
    )

    present_count = sum(1 for r, _ in records if r.status == AttendanceStatus.PRESENT)
    absent_count = sum(1 for r, _ in records if r.status == AttendanceStatus.ABSENT)

    return render_template(
        "reports/daily.html",
        title="Daily Report",
        records=records,
        report_date=report_date,
        present_count=present_count,
        absent_count=absent_count,
        AttendanceStatus=AttendanceStatus,
    )


@reports_bp.route("/student/<int:student_id>")
@login_required
def student_report(student_id: int):
    """Per-student detailed attendance history."""
    student = Student.query.get_or_404(student_id)

    records = (
        Attendance.query
        .filter_by(student_id=student_id)
        .order_by(Attendance.date.desc())
        .all()
    )

    total = len(records)
    present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
    absent = total - present
    percentage = round((present / total * 100), 1) if total > 0 else 0

    sms_records = SMSLog.query.filter_by(student_id=student_id).order_by(SMSLog.sent_time.desc()).all()

    return render_template(
        "reports/student.html",
        title=f"Report — {student.name}",
        student=student,
        records=records,
        present=present,
        absent=absent,
        percentage=percentage,
        sms_records=sms_records,
        AttendanceStatus=AttendanceStatus,
    )


@reports_bp.route("/export/excel")
@login_required
def export_excel():
    """Export today's attendance to an Excel workbook (.xlsx)."""
    date_str = request.args.get("date", str(date.today()))
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        report_date = date.today()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl is required for Excel export. Run: pip install openpyxl", "danger")
        return redirect(url_for("reports.daily"))

    records = (
        db.session.query(Attendance, Student)
        .join(Student, Attendance.student_id == Student.id)
        .filter(Attendance.date == report_date)
        .order_by(Student.department, Student.name)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {report_date}"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    present_fill = PatternFill(fill_type="solid", fgColor="D4EDDA")
    absent_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Attendance Report — {report_date.strftime('%A, %d %B %Y')}"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["#", "Student ID", "Name", "Department", "Semester", "Subject", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, (att, stu) in enumerate(records, 1):
        row = i + 2
        values = [i, stu.student_id, stu.name, stu.department, stu.semester, att.subject, att.status.value]
        fill = present_fill if att.status == AttendanceStatus.PRESENT else absent_fill
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [5, 15, 25, 22, 10, 25, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"attendance_{report_date}.xlsx",
    )


@reports_bp.route("/export/pdf")
@login_required
def export_pdf():
    """Export today's attendance to a PDF file."""
    date_str = request.args.get("date", str(date.today()))
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        report_date = date.today()

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        flash("reportlab is required for PDF export. Run: pip install reportlab", "danger")
        return redirect(url_for("reports.daily"))

    records = (
        db.session.query(Attendance, Student)
        .join(Student, Attendance.student_id == Student.id)
        .filter(Attendance.date == report_date)
        .order_by(Student.department, Student.name)
        .all()
    )

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1E3A5F"),
        spaceAfter=12, alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey,
        spaceAfter=20, alignment=TA_CENTER,
    )

    elements = [
        Paragraph("Student Attendance Report", title_style),
        Paragraph(report_date.strftime("%A, %d %B %Y"), sub_style),
    ]

    # Table data
    header = ["#", "Student ID", "Name", "Department", "Semester", "Subject", "Status"]
    data = [header]
    for i, (att, stu) in enumerate(records, 1):
        data.append([
            str(i), stu.student_id, stu.name, stu.department,
            str(stu.semester), att.subject, att.status.value,
        ])

    present_color = colors.HexColor("#D4EDDA")
    absent_color = colors.HexColor("#F8D7DA")
    header_color = colors.HexColor("#1E3A5F")

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT", (0, 0), (-1, -1), 22),
    ])

    # Colour status rows
    for i, (att, _) in enumerate(records, 1):
        color = present_color if att.status == AttendanceStatus.PRESENT else absent_color
        style.add("BACKGROUND", (0, i), (-1, i), color)

    col_widths = [1 * cm, 3 * cm, 5 * cm, 4.5 * cm, 2.5 * cm, 5 * cm, 2.5 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(style)

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"attendance_{report_date}.pdf",
    )


# ─────────────────────── Faculty Management (Admin) ──────────────────────────

@reports_bp.route("/faculty")
@login_required
def faculty_list():
    """Admin-only: list all faculty accounts."""
    if not current_user.is_admin:
        abort(403)
    faculty = Faculty.query.order_by(Faculty.name).all()
    form = AddFacultyForm()
    return render_template("reports/faculty.html", title="Faculty Management", faculty=faculty, form=form)


@reports_bp.route("/faculty/add", methods=["POST"])
@login_required
def add_faculty():
    """Admin-only: create a faculty account."""
    if not current_user.is_admin:
        abort(403)
    form = AddFacultyForm()
    if form.validate_on_submit():
        if Faculty.query.filter_by(username=form.username.data.strip()).first():
            flash("Username already exists.", "danger")
        else:
            f = Faculty(
                name=form.name.data.strip(),
                username=form.username.data.strip(),
                role=form.role.data,
            )
            f.set_password(form.password.data)
            db.session.add(f)
            db.session.commit()
            flash(f"Faculty '{f.name}' created.", "success")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{field}: {error}", "danger")
    return redirect(url_for("reports.faculty_list"))


@reports_bp.route("/faculty/<int:fid>/delete", methods=["POST"])
@login_required
def delete_faculty(fid: int):
    """Admin-only: delete a faculty account (cannot delete self)."""
    if not current_user.is_admin:
        abort(403)
    if fid == current_user.id:
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("reports.faculty_list"))
    f = Faculty.query.get_or_404(fid)
    name = f.name
    db.session.delete(f)
    db.session.commit()
    flash(f"Faculty '{name}' deleted.", "warning")
    return redirect(url_for("reports.faculty_list"))
